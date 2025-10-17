"""
Handle Extractor Script
Searches for terms in Column D and extracts corresponding values from Column W
Preserves data integrity by working directly with .xlsx files
"""

from openpyxl import load_workbook, Workbook
import sys

def extract_handles(input_file, search_terms, output_file):
    """
    Extract handles based on search terms in Column D
    
    Args:
        input_file: Path to input .xlsx file
        search_terms: List of terms to search for in Column D
        output_file: Path to output .xlsx file
    """
    
    print(f"\nLoading workbook: {input_file}")
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active
    
    # Create output workbook
    output_wb = Workbook()
    output_ws = output_wb.active
    
    # Column indices (1-based for openpyxl)
    col_d = 4   # Column D
    col_w = 23  # Column W
    
    # Get headers from first row
    header_d = ws.cell(row=1, column=col_d).value
    header_w = ws.cell(row=1, column=col_w).value
    
    # Write headers to output
    output_ws.cell(row=1, column=1, value=header_d)
    output_ws.cell(row=1, column=2, value=header_w)
    
    print(f"Headers: [{header_d}] and [{header_w}]")
    print(f"\nSearching for terms: {', '.join(search_terms)}")
    print(f"Searching in Column D...")
    
    # Track output row
    output_row = 2
    matches_found = 0
    
    # Iterate through all rows (starting from row 2 to skip header)
    # Using iter_rows for better compatibility with read_only mode
    for row in ws.iter_rows(min_row=2):
        cell_d_value = row[col_d - 1].value  # -1 because iter_rows uses 0-based indexing
        
        # Convert to string for comparison, handle None values
        if cell_d_value is not None:
            cell_d_str = str(cell_d_value).strip()
            
            # Check if any search term matches
            for term in search_terms:
                if term.lower() in cell_d_str.lower():
                    # Match found! Extract both Column D and Column W values
                    cell_w_value = row[col_w - 1].value  # -1 because iter_rows uses 0-based indexing
                    
                    # Write to output file
                    output_ws.cell(row=output_row, column=1, value=cell_d_value)
                    output_ws.cell(row=output_row, column=2, value=cell_w_value)
                    
                    output_row += 1
                    matches_found += 1
                    break  # Avoid duplicate entries if multiple terms match
    
    # Close input workbook
    wb.close()
    
    # Save output workbook
    output_wb.save(output_file)
    output_wb.close()
    
    print(f"\n✓ Found {matches_found} matches")
    print(f"✓ Results saved to: {output_file}")
    
    return matches_found


def main():
    """Main function to run the handle extractor"""
    
    print("="*60)
    print("HANDLE EXTRACTOR")
    print("="*60)
    
    # Input file (fixed)
    input_file = "data/robinsons-collections-raw.xlsx"
    
    # Get search terms from user
    print("\nEnter the term(s) to search for in Column D.")
    print("(For multiple terms, separate with commas)")
    user_input = input("\nSearch term(s): ").strip()
    
    if not user_input:
        print("Error: No search terms provided")
        sys.exit(1)
    
    # Parse search terms (split by comma and clean up)
    search_terms = [term.strip() for term in user_input.split(',') if term.strip()]
    
    if not search_terms:
        print("Error: No valid search terms provided")
        sys.exit(1)
    
    # Create output filename
    # Use the first term for the filename (or combine all if preferred)
    term_for_filename = search_terms[0] if len(search_terms) == 1 else '-'.join(search_terms)
    # Clean filename (remove special characters)
    term_for_filename = ''.join(c for c in term_for_filename if c.isalnum() or c in ['-', '_'])
    output_file = f"{term_for_filename}-handles.xlsx"
    
    # Extract handles
    try:
        matches = extract_handles(input_file, search_terms, output_file)
        
        if matches == 0:
            print("\n⚠ Warning: No matches found for the specified term(s)")
        else:
            print(f"\n✓ Successfully extracted {matches} handle(s)")
            
    except FileNotFoundError:
        print(f"\n✗ Error: Could not find input file '{input_file}'")
        print("Please ensure the file exists in the current directory")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error: {str(e)}")
        sys.exit(1)
    
    print("\n" + "="*60)


if __name__ == "__main__":
    main()

