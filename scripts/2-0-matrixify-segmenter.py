"""
Matrixify Segmenter Script
Matches handles from a reference file against the master products-raw.xlsx file
and extracts all matching rows while preserving data integrity
"""

from openpyxl import load_workbook, Workbook
import sys
import os
import unicodedata

def normalize_handle(handle):
    """
    Normalize handle for matching purposes ONLY (original data is preserved in output)
    
    This function creates a robust normalized version for matching by:
    1. Converting to lowercase
    2. Applying Unicode NFKD normalization (decomposes characters)
    3. Keeping ONLY ASCII alphanumeric characters and hyphens
    4. Removing ALL special symbols, accents, and Unicode variants
    
    This ensures handles match regardless of:
    - Special symbols (™, ®, ©, etc.)
    - Accented characters (é, ñ, ü, etc.)
    - Unicode encoding variations
    - Case differences
    
    Args:
        handle: The handle string to normalize
        
    Returns:
        Normalized handle string for matching (ASCII alphanumeric + hyphens only)
    """
    if not handle:
        return ""
    
    # Step 1: Convert to string and lowercase
    normalized = str(handle).lower()
    
    # Step 2: Remove problematic Unicode symbols that NFKD would convert to letters
    # ™ (U+2122) → "TM", ® (U+00AE) → "R", etc.
    # Remove these BEFORE NFKD to prevent them becoming ASCII letters
    problematic_symbols = [
        '\u2122',  # ™ TRADE MARK SIGN
        '\u00AE',  # ® REGISTERED SIGN  
        '\u00A9',  # © COPYRIGHT SIGN
        '\u2120',  # ℠ SERVICE MARK
    ]
    for symbol in problematic_symbols:
        normalized = normalized.replace(symbol, '')
    
    # Step 3: Apply NFKD normalization for accented characters
    # This decomposes é → e + ´, ñ → n + ˜, etc.
    normalized = unicodedata.normalize('NFKD', normalized)
    
    # Step 4: Keep ONLY ASCII alphanumeric and hyphens
    # This removes:
    # - All remaining Unicode special symbols and emoji
    # - All combining diacritics (the accent marks after decomposition)
    # - All non-ASCII characters
    # - All punctuation except hyphens
    result = []
    for char in normalized:
        # Only keep ASCII alphanumeric (a-z, 0-9) and hyphens
        if (char.isascii() and char.isalnum()) or char == '-':
            result.append(char)
    
    normalized = ''.join(result)
    
    # Step 5: Clean up multiple consecutive hyphens and strip
    while '--' in normalized:
        normalized = normalized.replace('--', '-')
    
    return normalized.strip('-').strip()

def extract_handles_from_reference(reference_file):
    """
    Extract unique handles from Column B of the reference file
    
    Args:
        reference_file: Path to the reference .xlsx file
        
    Returns:
        Tuple of (normalized_handles_set, handle_mapping_dict)
        - normalized_handles_set: Set of normalized handles for matching
        - handle_mapping_dict: Maps normalized → original handle (preserves ™, ®, © etc.)
    """
    print(f"\nLoading reference file: {reference_file}")
    wb = load_workbook(reference_file, read_only=True, data_only=True)
    ws = wb.active
    
    col_b = 2  # Column B
    normalized_handles = set()
    handle_mapping = {}  # normalized → original
    
    # Verify header (should be 'Product: Handle' in Column B)
    header_b = ws.cell(row=1, column=col_b).value
    print(f"Column B header: [{header_b}]")
    
    # Extract all handles from Column B (starting from row 2)
    for row in ws.iter_rows(min_row=2):
        handle_value = row[col_b - 1].value  # -1 for 0-based indexing
        
        if handle_value is not None:
            handle_str = str(handle_value).strip()
            if handle_str:  # Only add non-empty handles
                # Normalize the handle for matching purposes
                normalized = normalize_handle(handle_str)
                normalized_handles.add(normalized)
                # Map normalized → ORIGINAL handle (with special chars intact)
                handle_mapping[normalized] = handle_str
    
    wb.close()
    
    print(f"✓ Found {len(normalized_handles)} unique handle(s) in reference file")
    
    return normalized_handles, handle_mapping


def extract_matching_rows(master_file, handles, handle_mapping):
    """
    Extract all rows from master file that match the provided handles
    Uses Unicode normalization for matching while preserving reference handles
    
    Args:
        master_file: Path to the master products-raw.xlsx file
        handles: Set of NORMALIZED handles to match against
        handle_mapping: Dict mapping normalized → original reference handle
        
    Returns:
        Tuple of (matching_rows, header_row, total_rows_processed)
    """
    print(f"\nLoading master file: {master_file}")
    wb = load_workbook(master_file, read_only=True, data_only=True)
    ws = wb.active
    
    col_b = 2  # Column B (Product: Handle)
    col_b_idx = col_b - 1  # 0-based index
    
    # Get header row
    header_row = []
    for cell in ws[1]:
        header_row.append(cell.value)
    
    print(f"Master file has {len(header_row)} columns")
    print(f"Searching for matching handles in Column B...")
    print(f"Using Unicode normalization for robust matching...")
    
    matching_rows = []
    total_rows_processed = 0
    
    # Iterate through all rows (starting from row 2 to skip header)
    for row in ws.iter_rows(min_row=2):
        total_rows_processed += 1
        
        handle_value = row[col_b_idx].value
        
        if handle_value is not None:
            handle_str = str(handle_value).strip()
            
            # Normalize the handle from master file for comparison
            normalized_handle = normalize_handle(handle_str)
            
            # Check if this NORMALIZED handle matches any from our reference file
            if normalized_handle in handles:
                # Extract all cell values from this row
                row_data = [cell.value for cell in row]
                
                # CRITICAL: Replace Column B with the ORIGINAL reference handle
                # This preserves special characters (™, ®, ©) from the reference file
                original_handle = handle_mapping.get(normalized_handle, handle_str)
                row_data[col_b_idx] = original_handle
                
                matching_rows.append(row_data)
    
    wb.close()
    
    print(f"✓ Processed {total_rows_processed} rows from master file")
    print(f"✓ Found {len(matching_rows)} matching row(s)")
    print(f"✓ Replaced handles with original reference handles (preserving special chars)")
    
    return matching_rows, header_row, total_rows_processed


def write_output_file(output_file, header_row, matching_rows):
    """
    Write the matching rows to output file
    
    Args:
        output_file: Path to output .xlsx file
        header_row: List of header values
        matching_rows: List of rows (each row is a list of cell values)
    """
    print(f"\nWriting output file: {output_file}")
    
    output_wb = Workbook()
    output_ws = output_wb.active
    
    # Write header row
    for col_idx, header_value in enumerate(header_row, start=1):
        output_ws.cell(row=1, column=col_idx, value=header_value)
    
    # Write all matching rows
    for row_idx, row_data in enumerate(matching_rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            output_ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Save output file
    output_wb.save(output_file)
    output_wb.close()
    
    print(f"✓ Output file saved successfully")


def main():
    """Main function to run the matrixify segmenter"""
    
    print("="*60)
    print("MATRIXIFY SEGMENTER")
    print("="*60)
    
    # Master file (fixed location)
    master_file = "raw/products-raw.xlsx"
    
    # Get reference file from user
    print("\nThis script will extract products from the master file")
    print("based on handles found in your reference file.")
    print("\nEnter the path to your reference file:")
    print("(e.g., data/womens-socks.xlsx or output/my-file.xlsx)")
    reference_file = input("\nReference file path: ").strip()
    
    if not reference_file:
        print("\n✗ Error: No file path provided")
        sys.exit(1)
    
    # Remove quotes if user copied path with quotes
    reference_file = reference_file.strip('"').strip("'")
    
    # Check if reference file exists
    if not os.path.exists(reference_file):
        print(f"\n✗ Error: Reference file not found: {reference_file}")
        sys.exit(1)
    
    # Check if master file exists
    if not os.path.exists(master_file):
        print(f"\n✗ Error: Master file not found: {master_file}")
        print("Please ensure products-raw.xlsx exists in the raw/ directory")
        sys.exit(1)
    
    try:
        # Step 1: Extract handles from reference file
        handles, handle_mapping = extract_handles_from_reference(reference_file)
        
        if len(handles) == 0:
            print("\n⚠ Warning: No handles found in reference file")
            sys.exit(1)
        
        # Step 2: Extract matching rows from master file
        matching_rows, header_row, total_rows_processed = extract_matching_rows(master_file, handles, handle_mapping)
        
        if len(matching_rows) == 0:
            print("\n⚠ Warning: No matching rows found in master file")
            sys.exit(1)
        
        # Step 3: Create output filename
        # Extract base filename without extension
        base_name = os.path.splitext(os.path.basename(reference_file))[0]
        # Use the directory of the reference file for output
        reference_dir = os.path.dirname(reference_file)
        if reference_dir:
            output_file = os.path.join(reference_dir, f"{base_name}-extracted-products.xlsx")
        else:
            output_file = f"{base_name}-extracted-products.xlsx"
        
        # Step 4: Write output file
        write_output_file(output_file, header_row, matching_rows)
        
        # Print summary statistics
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        print(f"Reference file:       {reference_file}")
        print(f"Unique handles:       {len(handles)}")
        print(f"Master rows checked:  {total_rows_processed}")
        print(f"Matching rows found:  {len(matching_rows)}")
        print(f"Output file:          {output_file}")
        print("="*60)
        print("\n✓ Processing complete!")
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: File not found - {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

