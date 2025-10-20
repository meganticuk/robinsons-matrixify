"""
Size Tag Extractor Script
Extracts unique size tags from Column H of an Excel file
and outputs them to a sorted, deduplicated text file
"""

from openpyxl import load_workbook
import sys
import os

def extract_size_tags(file_path):
    """
    Extract all unique size tags from Column H of the Excel file
    
    Args:
        file_path: Path to the .xlsx file to analyze
        
    Returns:
        Sorted list of unique size tags
    """
    print(f"\nLoading file: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    col_h = 8  # Column H
    
    # Verify header
    header_h = ws.cell(row=1, column=col_h).value
    print(f"Column H header: [{header_h}]")
    
    size_tags = set()  # Use set for automatic deduplication
    total_rows_processed = 0
    rows_with_size_tags = 0
    
    # Iterate through all rows (starting from row 2 to skip header)
    print(f"\nProcessing rows...")
    for row in ws.iter_rows(min_row=2):
        total_rows_processed += 1
        
        # Get value from Column H (0-based index)
        cell_value = row[col_h - 1].value
        
        if cell_value is not None:
            # Convert to string and split by comma
            tags_string = str(cell_value).strip()
            
            if tags_string:
                # Split by comma and process each tag
                tags = [tag.strip() for tag in tags_string.split(',')]
                
                # Extract only tags that start with 'size_'
                size_tags_in_row = [tag for tag in tags if tag.startswith('size_')]
                
                if size_tags_in_row:
                    rows_with_size_tags += 1
                    # Add to our set (automatically deduplicates)
                    size_tags.update(size_tags_in_row)
    
    wb.close()
    
    print(f"✓ Processed {total_rows_processed} rows")
    print(f"✓ Found size tags in {rows_with_size_tags} rows")
    print(f"✓ Extracted {len(size_tags)} unique size tag(s)")
    
    # Convert set to sorted list
    sorted_size_tags = sorted(size_tags)
    
    return sorted_size_tags


def write_output_file(size_tags, output_file):
    """
    Write the size tags to a text file, one per line
    
    Args:
        size_tags: List of size tags to write
        output_file: Path to output text file
    """
    print(f"\nWriting output file: {output_file}")
    
    with open(output_file, 'w', encoding='utf-8') as f:
        for tag in size_tags:
            f.write(f"{tag}\n")
    
    print(f"✓ Output file saved successfully")


def main():
    """Main function to run the size tag extractor"""
    
    print("="*60)
    print("SIZE TAG EXTRACTOR")
    print("="*60)
    
    # Get file path from user
    print("\nThis script will extract all unique size tags from Column H")
    print("of your Excel file and save them to 'size_tags.txt'")
    print("\nEnter the full path to your .xlsx file:")
    print("(e.g., data/robinsons-socks.xlsx)")
    file_path = input("\nFile path: ").strip()
    
    if not file_path:
        print("\n✗ Error: No file path provided")
        sys.exit(1)
    
    # Remove quotes if user copied path with quotes
    file_path = file_path.strip('"').strip("'")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"\n✗ Error: File not found: {file_path}")
        sys.exit(1)
    
    # Check if it's an .xlsx file
    if not file_path.lower().endswith('.xlsx'):
        print(f"\n✗ Error: File must be an .xlsx file (got: {file_path})")
        print("We cannot use CSV files due to data integrity issues with ASCII conversion")
        sys.exit(1)
    
    try:
        # Step 1: Extract size tags
        size_tags = extract_size_tags(file_path)
        
        if len(size_tags) == 0:
            print("\n⚠ Warning: No size tags found in the file")
            print("Make sure Column H contains tags in the format: size_XXXXX")
            sys.exit(1)
        
        # Step 2: Create output filename in same directory as input file
        file_dir = os.path.dirname(file_path)
        if file_dir:
            output_file = os.path.join(file_dir, "size_tags.txt")
        else:
            output_file = "size_tags.txt"
        
        # Step 3: Write output file
        write_output_file(size_tags, output_file)
        
        # Print summary statistics
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        print(f"Input file:           {file_path}")
        print(f"Unique size tags:     {len(size_tags)}")
        print(f"Output file:          {output_file}")
        print("="*60)
        
        # Show first 10 size tags as preview
        print("\nFirst 10 size tags:")
        for i, tag in enumerate(size_tags[:10], 1):
            print(f"  {i}. {tag}")
        
        if len(size_tags) > 10:
            print(f"  ... and {len(size_tags) - 10} more")
        
        print("\n✓ Processing complete!")
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: File not found - {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

