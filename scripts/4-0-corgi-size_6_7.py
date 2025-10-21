"""
Gender Updater Script
This script updates gender tags in the products file based on brand and size criteria.

Features:
- Reads from raw/products-raw.xlsx (never modifies the original)
- Outputs ONLY matched rows to data/ directory with descriptive filename
- Output file contains only the rows that matched the filter criteria
- Exact case matching for brand name and size label
- Parses comma-separated tags in Column H
- Smart gender replacement logic:
  * Replaces opposite gender (Female -> Male or Male -> Female)
  * Keeps same gender unchanged
  * Appends to Unisex
- Shopify Matrixify format compliance:
  * Gender tag applied ONLY to first row of each unique handle
  * Subsequent variant rows have blank gender field
- Interactive prompts for brand, size, and gender
"""

import openpyxl
import json
from pathlib import Path


def parse_comma_separated_tags(cell_value):
    """
    Parse a cell value containing comma-separated tags.
    Returns a list of stripped tag values.
    """
    if cell_value is None or cell_value == "":
        return []
    
    # Split by comma and strip whitespace
    tags = [tag.strip() for tag in str(cell_value).split(',')]
    return [tag for tag in tags if tag]  # Remove empty strings


def parse_json_list(cell_value):
    """
    Parse a cell value that might be a JSON list string.
    Returns a list of values or empty list if parsing fails.
    """
    if cell_value is None or cell_value == "":
        return []
    
    # If it's already a list, return it
    if isinstance(cell_value, list):
        return cell_value
    
    # Try to parse as JSON
    try:
        parsed = json.loads(cell_value)
        if isinstance(parsed, list):
            return parsed
        else:
            return [parsed]
    except (json.JSONDecodeError, TypeError):
        # If not valid JSON, treat as plain string
        return [str(cell_value)]


def format_json_list(values):
    """
    Format a list of values as a JSON string.
    """
    return json.dumps(values)


def get_opposite_gender(gender):
    """
    Get the opposite gender for replacement logic.
    """
    gender_lower = gender.lower()
    if gender_lower == "male":
        return "Female"
    elif gender_lower == "female":
        return "Male"
    return None


def update_gender_list(current_genders, new_gender):
    """
    Update the gender list according to the business rules:
    - If new gender already exists, no change
    - If opposite gender exists, replace it with new gender
    - If "Unisex" exists, append new gender (if not already there)
    - Otherwise, just add the new gender
    
    Returns: (updated_list, was_changed)
    """
    # If new gender already in list, no change needed
    if new_gender in current_genders:
        return current_genders, False
    
    # Get opposite gender
    opposite = get_opposite_gender(new_gender)
    
    # Check if opposite gender exists
    if opposite and opposite in current_genders:
        # Replace opposite with new gender
        updated = [new_gender if g == opposite else g for g in current_genders]
        return updated, True
    
    # Check if Unisex exists
    if "Unisex" in current_genders:
        # Append new gender to Unisex
        return current_genders + [new_gender], True
    
    # If empty or no special cases, just add new gender
    if not current_genders:
        return [new_gender], True
    
    # Default: append
    return current_genders + [new_gender], True


def update_gender_tags():
    """
    Main function to update gender tags in the products file.
    """
    # File paths
    input_file = Path(r"C:\Users\New\Documents\Work\Client\Robinson\phase-2\raw\products-raw.xlsx")
    
    # Check if file exists
    if not input_file.exists():
        print(f"Error: File not found at {input_file}")
        return
    
    # Get user inputs
    print("=" * 60)
    print("Gender Tag Updater")
    print("=" * 60)
    print("\nThis script will update gender tags based on brand and size criteria.")
    print("Note: Exact case matching will be used for accuracy.\n")
    
    brand_name = input("Enter the Brand Name (exact case): ").strip()
    size_label = input("Enter the Size Label (exact case): ").strip()
    gender = input("Enter the Gender to add: ").strip()
    
    if not brand_name or not size_label or not gender:
        print("\nError: All fields are required!")
        return
    
    # Generate output filename
    output_dir = Path(r"C:\Users\New\Documents\Work\Client\Robinson\phase-2\data")
    output_file = output_dir / f"products-updated-{brand_name.lower()}-{size_label.replace('/', '-')}.xlsx"
    
    print(f"\nSearching for:")
    print(f"  Brand: '{brand_name}'")
    print(f"  Size: '{size_label}'")
    print(f"  Gender to add: '{gender}'")
    print(f"\nInput file: {input_file.name}")
    print(f"Output file: {output_file.name}")
    print("\nLoading workbook...")
    
    # Load the input workbook
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # Create new workbook for output (only matched rows)
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    
    # Copy header row to output
    header_row = list(ws_input[1])
    for col_idx, cell in enumerate(header_row, start=1):
        ws_output.cell(row=1, column=col_idx, value=cell.value)
    
    # Column indices (1-based for openpyxl)
    HANDLE_COL = 2  # Column B
    BRAND_COL = 6  # Column F
    SIZE_COL = 8   # Column H
    GENDER_COL = 95  # Column CQ (C=3, Q=17 -> 3*26+17 = 95)
    
    # Track statistics
    rows_processed = 0
    rows_matched = 0
    rows_updated = 0
    rows_unchanged = 0
    output_row = 2  # Start at row 2 (after header)
    
    # Track unique handles (for Shopify Matrixify format)
    seen_handles = set()  # Track which handles we've already output
    
    # Get total rows
    total_rows = ws_input.max_row
    
    print(f"Processing {total_rows} rows...")
    print()
    
    # Iterate through rows (skip header row)
    for row_num in range(2, total_rows + 1):
        rows_processed += 1
        
        # Show progress every 1000 rows
        if rows_processed % 1000 == 0:
            print(f"  Processed {rows_processed}/{total_rows - 1} rows... (Matched: {rows_matched}, Updated: {rows_updated})")
        
        # Get handle, brand, size, and gender cells from input
        handle_cell = ws_input.cell(row=row_num, column=HANDLE_COL)
        brand_cell = ws_input.cell(row=row_num, column=BRAND_COL)
        size_cell = ws_input.cell(row=row_num, column=SIZE_COL)
        gender_cell = ws_input.cell(row=row_num, column=GENDER_COL)
        
        handle_value = handle_cell.value
        brand_value = brand_cell.value
        size_tags = parse_comma_separated_tags(size_cell.value)
        
        # Check if both criteria match (exact case)
        # Brand must match exactly, and size_label must be in the comma-separated tags
        if brand_value == brand_name and size_label in size_tags:
            rows_matched += 1
            
            # Check if this is the first occurrence of this handle
            is_first_occurrence = handle_value not in seen_handles
            
            # Parse existing gender values (only if first occurrence)
            if is_first_occurrence:
                current_genders = parse_json_list(gender_cell.value)
                
                # Update gender list according to business rules
                updated_genders, was_changed = update_gender_list(current_genders, gender)
                
                if was_changed:
                    rows_updated += 1
                else:
                    rows_unchanged += 1
                
                # Mark this handle as seen
                seen_handles.add(handle_value)
            else:
                # Subsequent occurrence - no gender update needed
                updated_genders = None
            
            # Copy entire row to output workbook
            for col_idx in range(1, ws_input.max_column + 1):
                source_cell = ws_input.cell(row=row_num, column=col_idx)
                target_cell = ws_output.cell(row=output_row, column=col_idx)
                
                # Copy value (update gender if this column is the gender column AND it's first occurrence)
                if col_idx == GENDER_COL:
                    if is_first_occurrence:
                        target_cell.value = format_json_list(updated_genders)
                    else:
                        target_cell.value = None  # Leave blank for subsequent variants
                else:
                    target_cell.value = source_cell.value
            
            output_row += 1
    
    # Close input workbook (we don't modify it)
    wb_input.close()
    
    # Save the output workbook (only matched rows)
    print()
    print("Saving workbook...")
    wb_output.save(output_file)
    
    # Print summary
    print()
    print("=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"Total rows processed: {rows_processed}")
    print(f"Rows matched criteria: {rows_matched}")
    print(f"Rows updated: {rows_updated}")
    print(f"Rows unchanged (already correct): {rows_unchanged}")
    print()
    print("Update Rules Applied:")
    print(f"  - If gender was already '{gender}': No change")
    print(f"  - If gender was opposite: Replaced with '{gender}'")
    print(f"  - If gender was 'Unisex': Appended '{gender}'")
    print()
    print("Shopify Matrixify Format:")
    print(f"  - Gender tags added ONLY to first row of each unique handle")
    print(f"  - Subsequent variant rows have blank gender field")
    print(f"  - Unique handles processed: {len(seen_handles)}")
    print()
    print(f"Output file contains {rows_matched} matched rows (header + data)")
    print(f"Output saved to: {output_file}")
    print()
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    try:
        update_gender_tags()
    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nPress Enter to exit...")

