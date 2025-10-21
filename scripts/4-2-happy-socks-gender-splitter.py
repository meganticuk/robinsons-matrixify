"""
Happy Socks Gender Splitter Script
Processes ALL Happy Socks products and splits them into 3 files based on size criteria.

Logic:
- Products with ONLY size_36_40 → Female → Output to: happy-socks-female-only.xlsx
- Products with ONLY size_41_46 → Male → Output to: happy-socks-male-only.xlsx  
- Products with BOTH sizes → Female + Male + Unisex → Output to: happy-socks-unisex.xlsx

Features:
- Reads from all-socks-preupload-extracted-products.xlsx
- Fixed brand: "Happy Socks"
- Creates 3 separate output files in data/ directory
- Applies appropriate gender tags to each category
- Shopify Matrixify format compliance
- All processing in one pass for efficiency
"""

import openpyxl
import json
from pathlib import Path


def parse_comma_separated_tags(cell_value):
    """
    Parse a cell value containing comma-separated tags.
    Returns a list of stripped tag values.
    """
    if cell_value is None or cell_value == "":
        return []
    
    # Split by comma and strip whitespace
    tags = [tag.strip() for tag in str(cell_value).split(',')]
    return [tag for tag in tags if tag]  # Remove empty strings


def format_json_list(values):
    """
    Format a list of values as a JSON string.
    """
    return json.dumps(values)


def categorize_product(size_tags):
    """
    Categorize a product based on which sizes it has.
    
    Returns: tuple (category, gender_list)
    - category: "female_only", "male_only", or "unisex"
    - gender_list: list of gender strings to apply
    """
    has_36_40 = "size_36_40" in size_tags
    has_41_46 = "size_41_46" in size_tags
    
    if has_36_40 and has_41_46:
        # Has BOTH sizes → Unisex
        return "unisex", ["Female", "Male", "Unisex"]
    elif has_36_40:
        # Has ONLY size_36_40 → Female
        return "female_only", ["Female"]
    elif has_41_46:
        # Has ONLY size_41_46 → Male
        return "male_only", ["Male"]
    else:
        # Has neither size → skip
        return None, []


def split_happy_socks():
    """
    Main function to split Happy Socks products into 3 files.
    """
    print("=" * 60)
    print("Happy Socks Gender Splitter")
    print("=" * 60)
    print("\nThis script processes all Happy Socks products and splits them into:")
    print("  1. Female-only (size_36_40 only)")
    print("  2. Male-only (size_41_46 only)")
    print("  3. Unisex (both sizes)")
    print()
    
    # Fixed paths
    input_file = Path(r"C:\Users\New\Documents\Work\Client\Robinson\phase-2\data\all-socks-preupload-extracted-products.xlsx")
    output_dir = Path(r"C:\Users\New\Documents\Work\Client\Robinson\phase-2\data")
    
    # Check if input file exists
    if not input_file.exists():
        print(f"Error: File not found at {input_file}")
        return
    
    # Output files
    output_files = {
        "female_only": output_dir / "happy-socks-female-only.xlsx",
        "male_only": output_dir / "happy-socks-male-only.xlsx",
        "unisex": output_dir / "happy-socks-unisex.xlsx"
    }
    
    print(f"Input file: {input_file.name}")
    print(f"\nOutput files:")
    print(f"  Female-only: {output_files['female_only'].name}")
    print(f"  Male-only: {output_files['male_only'].name}")
    print(f"  Unisex: {output_files['unisex'].name}")
    print("\nLoading workbook...")
    
    # Load the input workbook
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # Create output workbooks
    wb_outputs = {
        "female_only": openpyxl.Workbook(),
        "male_only": openpyxl.Workbook(),
        "unisex": openpyxl.Workbook()
    }
    
    # Get header row from input
    header_row = list(ws_input[1])
    
    # Copy header row to all output workbooks and initialize
    for category, wb in wb_outputs.items():
        ws = wb.active
        for col_idx, cell in enumerate(header_row, start=1):
            ws.cell(row=1, column=col_idx, value=cell.value)
    
    # Column indices (1-based for openpyxl)
    HANDLE_COL = 2   # Column B
    BRAND_COL = 6    # Column F
    SIZE_COL = 8     # Column H
    GENDER_COL = 95  # Column CQ
    
    # Track statistics
    rows_processed = 0
    happy_socks_found = 0
    categorized = {
        "female_only": 0,
        "male_only": 0,
        "unisex": 0,
        "skipped": 0  # Happy Socks without relevant sizes
    }
    
    # Track unique handles per category (for Shopify Matrixify format)
    seen_handles = {
        "female_only": set(),
        "male_only": set(),
        "unisex": set()
    }
    
    # Track output row numbers for each workbook
    output_rows = {
        "female_only": 2,
        "male_only": 2,
        "unisex": 2
    }
    
    # Get total rows
    total_rows = ws_input.max_row
    
    print(f"Processing {total_rows} rows...")
    print()
    
    # Iterate through rows (skip header row)
    for row_num in range(2, total_rows + 1):
        rows_processed += 1
        
        # Show progress every 500 rows
        if rows_processed % 500 == 0:
            print(f"  Processed {rows_processed}/{total_rows - 1} rows... (Happy Socks: {happy_socks_found})")
        
        # Get cells from input
        handle_cell = ws_input.cell(row=row_num, column=HANDLE_COL)
        brand_cell = ws_input.cell(row=row_num, column=BRAND_COL)
        size_cell = ws_input.cell(row=row_num, column=SIZE_COL)
        
        handle_value = handle_cell.value
        brand_value = brand_cell.value
        size_tags = parse_comma_separated_tags(size_cell.value)
        
        # Only process Happy Socks products
        if brand_value != "Happy Socks":
            continue
        
        happy_socks_found += 1
        
        # Categorize the product
        category, gender_list = categorize_product(size_tags)
        
        if category is None:
            # Happy Socks but doesn't have the sizes we care about
            categorized["skipped"] += 1
            continue
        
        # Track if this is the first occurrence of this handle in this category
        is_first_occurrence = handle_value not in seen_handles[category]
        
        if is_first_occurrence:
            seen_handles[category].add(handle_value)
            categorized[category] += 1
        
        # Get the output workbook and row number for this category
        ws_output = wb_outputs[category].active
        output_row = output_rows[category]
        
        # Copy entire row to output workbook
        for col_idx in range(1, ws_input.max_column + 1):
            source_cell = ws_input.cell(row=row_num, column=col_idx)
            target_cell = ws_output.cell(row=output_row, column=col_idx)
            
            # Copy value (update gender if this is the gender column AND first occurrence)
            if col_idx == GENDER_COL:
                if is_first_occurrence:
                    target_cell.value = format_json_list(gender_list)
                else:
                    target_cell.value = None  # Leave blank for subsequent variants
            else:
                target_cell.value = source_cell.value
        
        # Increment output row for this category
        output_rows[category] += 1
    
    # Close input workbook
    wb_input.close()
    
    # Save all output workbooks
    print()
    print("Saving workbooks...")
    for category, wb in wb_outputs.items():
        wb.save(output_files[category])
        print(f"  Saved {output_files[category].name}")
    
    # Print summary
    print()
    print("=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"Total rows processed: {rows_processed}")
    print(f"Happy Socks products found: {happy_socks_found}")
    print()
    print("Categorization:")
    print(f"  Female-only (size_36_40 only):")
    print(f"    - Unique products: {len(seen_handles['female_only'])}")
    print(f"    - Total rows: {output_rows['female_only'] - 2}")
    print(f"    - Gender tags: ['Female']")
    print()
    print(f"  Male-only (size_41_46 only):")
    print(f"    - Unique products: {len(seen_handles['male_only'])}")
    print(f"    - Total rows: {output_rows['male_only'] - 2}")
    print(f"    - Gender tags: ['Male']")
    print()
    print(f"  Unisex (both sizes):")
    print(f"    - Unique products: {len(seen_handles['unisex'])}")
    print(f"    - Total rows: {output_rows['unisex'] - 2}")
    print(f"    - Gender tags: ['Female', 'Male', 'Unisex']")
    print()
    print(f"  Skipped (no relevant sizes): {categorized['skipped']}")
    print()
    print("Shopify Matrixify Format:")
    print("  - Gender tags applied ONLY to first row of each unique handle")
    print("  - Subsequent variant rows have blank gender field")
    print()
    print("Output files saved to data/ directory!")
    print()
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    try:
        split_happy_socks()
    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nPress Enter to exit...")

