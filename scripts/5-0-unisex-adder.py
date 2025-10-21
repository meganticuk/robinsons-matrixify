"""
Unisex Gender Tag Adder
This script adds "Unisex" to products that have BOTH "Male" AND "Female" gender tags.

Features:
- Processes ALL products in the input file
- If a product has both "Male" and "Female" → adds "Unisex" (if not already present)
- Preserves all other data
- Shopify Matrixify format compliance:
  * Gender tag applied ONLY to first row of each unique handle
  * Subsequent variant rows have blank gender field
- Interactive prompt for input file path
- Outputs to data/ directory with descriptive filename

Logic:
- ["Male", "Female"] → ["Male", "Female", "Unisex"]
- ["Female", "Male"] → ["Female", "Male", "Unisex"]
- ["Male", "Female", "Unisex"] → No change (already has Unisex)
- ["Male"] → No change
- ["Female"] → No change
- ["Unisex"] → No change
"""

import openpyxl
import json
from pathlib import Path


def parse_json_list(cell_value):
    """
    Parse a cell value that might be a JSON list string.
    Returns a list of values or empty list if parsing fails.
    """
    if cell_value is None or cell_value == "":
        return []
    
    # If it's already a list, return it
    if isinstance(cell_value, list):
        return cell_value
    
    # Try to parse as JSON
    try:
        parsed = json.loads(cell_value)
        if isinstance(parsed, list):
            return parsed
        else:
            return [parsed]
    except (json.JSONDecodeError, TypeError):
        # If not valid JSON, treat as plain string
        return [str(cell_value)]


def format_json_list(values):
    """
    Format a list of values as a JSON string.
    """
    return json.dumps(values)


def should_add_unisex(genders):
    """
    Check if "Unisex" should be added to the gender list.
    
    Returns True if:
    - List contains BOTH "Male" AND "Female"
    - List does NOT already contain "Unisex"
    
    Args:
        genders: List of gender strings
        
    Returns:
        Boolean - True if Unisex should be added
    """
    if not genders or len(genders) < 2:
        return False
    
    # Check if both Male and Female exist (case-insensitive)
    has_male = any(g.lower() == "male" for g in genders)
    has_female = any(g.lower() == "female" for g in genders)
    has_unisex = any(g.lower() == "unisex" for g in genders)
    
    # Add Unisex if we have both Male and Female but not Unisex yet
    return has_male and has_female and not has_unisex


def add_unisex_tags():
    """
    Main function to add Unisex tags to products with both Male and Female.
    """
    print("=" * 60)
    print("Unisex Gender Tag Adder")
    print("=" * 60)
    print("\nThis script adds 'Unisex' to products that have BOTH 'Male' AND 'Female'.")
    print("Products with only one gender or already having 'Unisex' are unchanged.\n")
    
    # Get input file path
    print("Enter the input file path:")
    print("  Example: data/products-updated-happy-socks-size_41_46.xlsx")
    input_file_str = input("\nInput file path: ").strip()
    
    if not input_file_str:
        print("\nError: Input file path is required!")
        return
    
    # Remove quotes if user copied path with quotes
    input_file_str = input_file_str.strip('"').strip("'")
    input_file = Path(input_file_str)
    
    # Check if file exists
    if not input_file.exists():
        print(f"\nError: File not found at {input_file}")
        return
    
    # Generate output filename
    output_dir = Path(r"C:\Users\New\Documents\Work\Client\Robinson\phase-2\data")
    input_basename = input_file.stem  # filename without extension
    output_file = output_dir / f"{input_basename}-with-unisex.xlsx"
    
    print(f"\nInput file: {input_file.name}")
    print(f"Output file: {output_file.name}")
    print("\nLoading workbook...")
    
    # Load the input workbook
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # Create new workbook for output
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    
    # Copy header row to output
    header_row = list(ws_input[1])
    for col_idx, cell in enumerate(header_row, start=1):
        ws_output.cell(row=1, column=col_idx, value=cell.value)
    
    # Column indices (1-based for openpyxl)
    HANDLE_COL = 2   # Column B
    GENDER_COL = 95  # Column CQ (C=3, Q=17 -> 3*26+17 = 95)
    
    # Track statistics
    rows_processed = 0
    rows_with_unisex_added = 0
    rows_unchanged = 0
    output_row = 2  # Start at row 2 (after header)
    
    # Track unique handles (for Shopify Matrixify format)
    seen_handles = set()  # Track which handles we've already processed
    
    # Get total rows
    total_rows = ws_input.max_row
    
    print(f"Processing {total_rows} rows...")
    print()
    
    # Iterate through rows (skip header row)
    for row_num in range(2, total_rows + 1):
        rows_processed += 1
        
        # Show progress every 1000 rows
        if rows_processed % 1000 == 0:
            print(f"  Processed {rows_processed}/{total_rows - 1} rows... (Unisex added: {rows_with_unisex_added})")
        
        # Get handle and gender cells from input
        handle_cell = ws_input.cell(row=row_num, column=HANDLE_COL)
        gender_cell = ws_input.cell(row=row_num, column=GENDER_COL)
        
        handle_value = handle_cell.value
        
        # Check if this is the first occurrence of this handle
        is_first_occurrence = handle_value not in seen_handles
        
        # Process gender (only if first occurrence)
        if is_first_occurrence:
            current_genders = parse_json_list(gender_cell.value)
            
            # Check if we should add Unisex
            if should_add_unisex(current_genders):
                updated_genders = current_genders + ["Unisex"]
                rows_with_unisex_added += 1
            else:
                updated_genders = current_genders
                rows_unchanged += 1
            
            # Mark this handle as seen
            seen_handles.add(handle_value)
        else:
            # Subsequent occurrence - no gender update needed
            updated_genders = None
        
        # Copy entire row to output workbook
        for col_idx in range(1, ws_input.max_column + 1):
            source_cell = ws_input.cell(row=row_num, column=col_idx)
            target_cell = ws_output.cell(row=output_row, column=col_idx)
            
            # Copy value (update gender if this column is the gender column AND it's first occurrence)
            if col_idx == GENDER_COL:
                if is_first_occurrence and updated_genders is not None:
                    target_cell.value = format_json_list(updated_genders) if updated_genders else None
                else:
                    target_cell.value = None  # Leave blank for subsequent variants
            else:
                target_cell.value = source_cell.value
        
        output_row += 1
    
    # Close input workbook
    wb_input.close()
    
    # Save the output workbook
    print()
    print("Saving workbook...")
    wb_output.save(output_file)
    
    # Print summary
    print()
    print("=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"Total rows processed: {rows_processed}")
    print(f"Unique handles processed: {len(seen_handles)}")
    print(f"Rows with Unisex added: {rows_with_unisex_added}")
    print(f"Rows unchanged: {rows_unchanged}")
    print()
    print("Logic Applied:")
    print("  - Products with BOTH 'Male' AND 'Female' → Added 'Unisex'")
    print("  - Products already having 'Unisex' → No change")
    print("  - Products with only one gender → No change")
    print()
    print("Shopify Matrixify Format:")
    print("  - Gender tags applied ONLY to first row of each unique handle")
    print("  - Subsequent variant rows have blank gender field")
    print()
    print(f"Output saved to: {output_file}")
    print()
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    try:
        add_unisex_tags()
    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nPress Enter to exit...")

